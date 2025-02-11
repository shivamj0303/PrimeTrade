import requests
import json
import pandas as pd
import time
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF

# CoinGecko API URL
API_URL = "https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&order=market_cap_desc&per_page=50&page=1&sparkline=false"

# Excel file path
EXCEL_FILE = "crypto_data.xlsx"

def fetch_crypto_data():
    try:
        response = requests.get(API_URL)
        response.raise_for_status()  # Raise an exception for bad status codes (4xx or 5xx)
        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data: {e}")
        return None

def analyze_data(data):
    if data is None:
        return None

    df = pd.DataFrame(data)
    
    # Selecting required fields
    df = df[['name', 'symbol', 'current_price', 'market_cap', 'total_volume', 'price_change_percentage_24h']]
    df.rename(columns={'current_price': 'Price (USD)', 'market_cap': 'Market Cap', 'total_volume': '24h Volume', 'price_change_percentage_24h': '24h Change (%)'}, inplace=True)


    top_5 = df.head(5)
    average_price = df['Price (USD)'].mean()
    highest_change = df['24h Change (%)'].max()
    lowest_change = df['24h Change (%)'].min()

    analysis = {
        "top_5": top_5.to_dict('records'),  # Convert to list of dictionaries for easier handling
        "average_price": average_price,
        "highest_change": highest_change,
        "lowest_change": lowest_change
    }
    return df, analysis


def update_excel(df):
    try:
        if os.path.exists(EXCEL_FILE):
            book = load_workbook(EXCEL_FILE)
            sheet = book.active  # Assumes data is in the active sheet. Change if needed.

            # Clear existing data (from row 2 onwards, assuming header is in row 1)
            sheet.delete_rows(2, sheet.max_row)

            for row in dataframe_to_rows(df, header=False, index=False):  # Append data without header
                sheet.append(row)

            book.save(EXCEL_FILE)
            print("Excel sheet updated.")
        else:
            df.to_excel(EXCEL_FILE, index=False)  # Create new sheet if it doesn't exist
            print("Excel sheet created.")

    except Exception as e:
        print(f"Error updating Excel: {e}")


def generate_pdf_report(analysis, timestamp):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Crypto Data Report - {timestamp}", ln=1, align="C")

    pdf.set_font("Arial", size=10)
    pdf.cell(200, 10, txt=f"Top 5 Cryptocurrencies by Market Cap:", ln=1)

    for coin in analysis['top_5']:
        pdf.cell(200, 5, txt=f"{coin['name']} ({coin['symbol']}): ${coin['Price (USD)']:.2f}, Market Cap: ${coin['Market Cap']:.2f}, 24h Change: {coin['24h Change (%)']:.2f}%", ln=1)

    pdf.cell(200, 10, txt=f"\nAverage Price (Top 50): ${analysis['average_price']:.2f}", ln=1)
    pdf.cell(200, 10, txt=f"Highest 24h Change: {analysis['highest_change']:.2f}%", ln=1)
    pdf.cell(200, 10, txt=f"Lowest 24h Change: {analysis['lowest_change']:.2f}%", ln=1)


    pdf_filename = f"crypto_report_{timestamp.replace(':', '-').replace(' ', '_')}.pdf"  # Sanitize filename
    pdf.output(pdf_filename)
    print(f"PDF report generated: {pdf_filename}")


def main():
    while True:
        data = fetch_crypto_data()
        if data:
            df, analysis = analyze_data(data)
            if df is not None and analysis is not None:
                update_excel(df)
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                generate_pdf_report(analysis, timestamp)
        
        time.sleep(10)  # Wait for 5 minutes (300 seconds)


if __name__ == "__main__":
    main()